# Importa estilos de formatação da biblioteca openpyxl
from openpyxl.styles import Font, PatternFill  # Font (fonte do texto) e PatternFill (preenchimento de células)
from openpyxl.styles import Border, Side  # Border (borda das células) e Side (estilo das bordas)
from openpyxl.styles import Alignment  # Alignment (alinhamento do texto dentro das células)

# Importa funções para manipulação de arquivos Excel
from openpyxl import load_workbook  # Carrega um arquivo Excel existente
from openpyxl import Workbook  # Cria um novo arquivo Excel

# Importa a classe datetime para trabalhar com datas e horários
from datetime import datetime  

# Importa a classe Path do módulo pathlib para manipulação de caminhos de arquivos
from pathlib import Path  

# Importa o módulo unicodedata para manipulação de caracteres Unicode (como remoção de acentos)
import unicodedata  

# Importa o módulo tempfile para criar arquivos temporários no sistema
import tempfile  

# Importa o módulo os para interagir com o sistema operacional (como manipulação de arquivos e diretórios)
import os  



def len_file_path(path, mode=0):
    # Essa função recebe como parâmetro o caminho de uma pasta (path) e um modo (mode)
    # Ele verifica a quantidade de arquivos que existem dentro da pasta (path)
    # O mode, serve para definir qual arquivo vai ser contabilizado
    # mode = 0, vai verificar a quantidade de arquivos DATABASE (.xlsx) existem e vai retornar a quantidade atual
    # mode = 1, vai verificar a quantidade de arquivos GRÁFICO-BARRA existem e vai retornar a quantidade atual
    # mode = 2, vai verificar a quantidade de arquivos GRÁFICO-LINHA existem e vai retornar a quantidade atual
    # mode = 3, vai verificar a quantidade de arquivos GRÁFICO-TEMPERATURA-NUMÉRICA existem e vai retornar a quantidade atual
    # mode = 4, vai verificar a quantidade de arquivos GRÁFICO-TEMPERATURA-ANALÍTICA existem e vai retornar a quantidade atual
    counter = 0
    if mode == 0:
        for file in os.listdir(path):
            if 'DATABASE_' in file:
                counter += 1
    elif mode == 1:
        for file in os.listdir(path):
            if remover_utf8('GRÁFICO-BARRA_') in file:
                counter += 1
    elif mode == 2:
        for file in os.listdir(path):
            if remover_utf8('GRÁFICO-LINHA_') in file:
                counter += 1
    elif mode == 3:
        for file in os.listdir(path):
            if remover_utf8('GRÁFICO-TEMPERATURA-NUMÉRICA_') in file:
                counter += 1
    elif mode == 4:
        for file in os.listdir(path):
            if remover_utf8('GRÁFICO-TEMPERATURA-ANALÍTICA_') in file:
                counter += 1
                
    return counter

def generate_name_file(files=int, name=str, extension=str):
    # Essa função serve para gerar um nome para o arquivo, seja lá ele qual for
    
    # Essa função vai receber a quantidade de arquivos atuais que existem (files), com um determinado nome (name)
    # Quando receber a quantidade (files), será obviamente um número inteiro, ou seja 1, 2, 3, ..., 1000, ..., 12123...
    # Eu quero transformar esse número num determinado padrão, que é o padrão de pelo menos 6 digitos
    # Ou seja, se eu receber o número 0, seria 000001, se receber o número 1, 000002, e assim por diante
    # Pois, files, é o número de arquivos atuais, como estou gerando mais um, quero adicionar uma nova numeração a ele
    
    # O extension, vai receber a extensão do arquivo que estou gerando o nome e vai agregar ao valor gerado anteriormente
    # Por exemplo, files=0, name='imagem', extension='png', o nome do meu arquivo vai ficar: 'imagem_000001.png' 
    extension = extension.replace('.', '')
    if files + 1 < 10:
        files = f'00000{files + 1}'
    elif files + 1 < 100:
        files = f'0000{files + 1}'
    elif files + 1 < 1000:
        files = f'000{files + 1}'
    elif files + 1 < 10000:
        files = f'00{files + 1}'
    elif files + 1 < 100000:
        files = f'0{files + 1}'
    else:
        files = f'{files + 1}'
    return f'{name}_{files}.{extension}'

def get_local_default():
    # A função obtem o local atual deste arquivo, e retorna o local em que o arquivo está salvo
    local_default = Path(__file__).parent.resolve()
    return local_default

def again_path(path):
    # Esta função recebe um local e retorna o mesmo local, só que um termo antes
    # Exemplo:
    # path='C:User/default/comment/jesus/src/generate_file.py'
    # Vai retornar='C:User/default/comment/jesus/src/'
    
    # path='C:User/default/comment/jesus/src/'
    # Vai retornar='C:User/default/comment/jesus/'
    
    return os.path.dirname(path)

def select_next_path(path, next):
    # A função recebe o local de uma pasta (path), e um arquivo ou pasta (next)
    # Recebe esses dois parâmetros e retorna um caminmho novo de adição entre path e next
    # Exemplo:
    # path='C:User/default/comment/jesus/src', next='banana.png'
    # Vai retornar='C:User/default/comment/jesus/src/banana.png'
    
    # path='C:User/default/comment/jesus/src', next='pasta'
    # Vai retornar='C:User/default/comment/jesus/src/pasta/'
    return os.path.join(path, next)

def remover_utf8(texto):
    # A função recebe um texto e remove toda e qualquer acentuação
    normalized = unicodedata.normalize('NFD', texto)
    new_text = ''.join(c for c in normalized if unicodedata.category(c) != 'Mn')
    return new_text

def settings_font():
    # A função serve para retornar configurações de borda, font e fundo de uma planilha excel (.xlsx)
    font_title = Font(color="ffffff", size=14, bold=True) # Cria uma configuração de fonte na cor branca, tamanho 14, em negrito
    fill_title = PatternFill(start_color="cc0000", end_color="cc0000", fill_type="solid") # Cria uma cor de fundo vermelha sólida, sem gradiente de cor
    font_display = Font(color="000000", size=12, bold=True) # Cria uma configuração de fonte na cor branca, tamanho 12, em negrito
    fill_display = PatternFill(start_color="ffea00", end_color="ffea00", fill_type="solid") # Cria uma cor de fundo amarela sólida, sem gradiente de cor
    
    border_all_min = Border(
        left=Side(border_style="thin"), # Borda esquerda fina
        right=Side(border_style="thin"), # Borda direita fina
        top=Side(border_style="thin"), # Borda superior fina
        bottom=Side(border_style="thin"), # Borda inferior fina
    )
    border_left_bottom = Border(
        left=Side(border_style="thick"), # Borda esquerda grossa
        right=Side(border_style="thin"), # Borda direita fina
        top=Side(border_style="thin"), # Borda superior fina
        bottom=Side(border_style="thick"), # Borda inferior grossa
    )
    border_right_bottom = Border(
        left=Side(border_style="thin"), # Borda esquerda fina
        right=Side(border_style="thick"), # Borda direita grossa
        top=Side(border_style="thin"), # Borda superior fina
        bottom=Side(border_style="thick"), # Borda inferior grossa
    )
    border_bottom = Border(
        left=Side(border_style="thin"), # Borda esquerda fina
        right=Side(border_style="thin"), # Borda direita fina
        top=Side(border_style="thin"), # Borda superior fina
        bottom=Side(border_style="thick"), # Borda inferior grossa
    )

    return [
        font_title,                # 00
        fill_title,                # 01
        font_display,              # 02
        fill_display,              # 03
       
        border_all_min,            # 04
        border_left_bottom,        # 05
        border_right_bottom,       # 06
        border_bottom,             # 07
        
    ]

def settings_border():
    border_1 = Border(
        top=Side(border_style='thin', color='000000'),
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000')
    )
    border_2 = Border(
        bottom=Side(border_style='thin', color='000000'),
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000')
    )
    border_3 = Border(left=Side(border_style='thin', color='000000'))
    border_4 = Border(right=Side(border_style='thin', color='000000'))
    return [
        border_1,
        border_2,
        border_3,
        border_4,
    ]

def settings_align():
    align_1 = Alignment(horizontal="center", vertical="center")
    align_2 = Alignment(horizontal="left", vertical="center")
    return [
        align_1,
        align_2,
    ]

def settings_header(header_1, header_2, date):
    header = [
            f"{header_1} - {date}".upper(),
            f"{header_2}".upper(),
    ]
    return header

def create_excel(array=list, path=str, name=str, extension=str):
    try:
        wb = Workbook()
        ws = wb.active

        ws.merge_cells("A1:D1")
        ws.merge_cells("A2:D2")
        
        date = datetime.now()
        date = f"{date.day}/{date.month}"
        
        header = settings_header('TRANSFERÊNCIA DE CALOR', 'JESUS', date)
        border_settings = settings_border()
        align_settings = settings_align()
        font_settings = settings_font()
        
        ws["A1"] = header[0]
        ws["A1"].alignment = align_settings[0]
        ws["A1"].font = font_settings[0]
        ws["A1"].fill = font_settings[1]
        ws["A1"].border = border_settings[0]

        ws["A2"] = header[1]
        ws["A2"].alignment = align_settings[0]
        ws["A2"].font = font_settings[0]
        ws["A2"].fill = font_settings[1]
        ws["A2"].border = border_settings[1]

        ws["A3"] = "PONTO Nº"
        ws["A3"].alignment = align_settings[0]
        ws["A3"].border = border_settings[2]

        ws["B3"] = "SOLUÇÃO NUMÉRICA"
        ws["B3"].alignment = align_settings[0]
        ws["B3"].border = font_settings[4]

        ws["C3"] = "SOLUÇÃO ANALÍTICA"
        ws["C3"].alignment = align_settings[0]
        ws["C3"].border = font_settings[4]

        ws["D3"] = "ERRO"
        ws["D3"].alignment = align_settings[0]
        ws["D3"].border = font_settings[4]

        for col in range(1, 5):
            ws.cell(row=3, column=col).font = font_settings[2]
            ws.cell(row=3, column=col).fill = font_settings[3]
            
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        
        counter = 0
        for c, data in enumerate(array):
            counter += 1
            calc_numeric = '{:0.20f}'.format(data[0])
            calc_analytic = '{:0.20f}'.format(data[1])
            error = '{:0.20f}'.format(data[2])
            
            ws.cell(row=counter + 3, column=1).value = counter
            ws.cell(row=counter + 3, column=1).alignment = align_settings[0]
            ws.cell(row=counter + 3, column=1).border = font_settings[4]

            ws.cell(row=counter + 3, column=2).value = calc_numeric
            ws.cell(row=counter + 3, column=2).alignment = align_settings[0]
            ws.cell(row=counter + 3, column=2).border = font_settings[4]

            ws.cell(row=counter + 3, column=3).value = calc_analytic
            ws.cell(row=counter + 3, column=3).alignment = align_settings[1]
            ws.cell(row=counter + 3, column=3).border = font_settings[4]

            ws.cell(row=counter + 3, column=4).value = error
            ws.cell(row=counter + 3, column=4).alignment = align_settings[0]
            ws.cell(row=counter + 3, column=4).border = font_settings[4]

        last_column = 5
        for col in range(1, last_column):
            ws.cell(row=len(array) + 3, column=col).border = font_settings[7]
            
        
        extension = extension.replace('.', '')
        extension = f'.{extension}'
        
        name_file = name.replace(extension, '')
        name_file = remover_utf8(name_file).upper()
        name_file = f'{name_file}{extension}'
        
        local_file = select_next_path(path, name_file)

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            wb.save(temp_file.name)  

        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as final_temp_file:
            workbook = load_workbook(temp_file.name)
            sheet = workbook.active

            sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
            sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE

            workbook.save(local_file)
        
        return 'Database salvo com sucesso!!!'
    except Exception as e:
        return 'Ocorreu um erro durante o salvamento...\n\n{e}'

  
